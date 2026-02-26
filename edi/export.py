"""
edi/export.py — Генерация печатных форм: XLSX, XML, PDF

Использование:
    from edi.export import export_xlsx, export_pdf, export_xml
    response = export_xlsx(documents, title="Заказы за февраль")
    response = export_pdf(documents)
    response = export_xml(documents)
"""
import io
from datetime import date
from django.http import HttpResponse

# ═══════════════════════════════════════════════════════
# XLSX
# ═══════════════════════════════════════════════════════

def export_xlsx(documents, title="Экспорт документов"):
    """Генерирует XLSX с листами: Сводка + отдельный лист на каждый тип."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl не установлен. Выполните: pip install openpyxl", status=500)

    wb = openpyxl.Workbook()

    # ── Стили ──
    HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TITLE_FONT   = Font(name="Arial", bold=True, size=13, color="1E3A5F")
    LABEL_FONT   = Font(name="Arial", bold=True, size=9, color="444444")
    DATA_FONT    = Font(name="Arial", size=9)
    ALT_FILL     = PatternFill("solid", fgColor="F0F4FA")
    TOTAL_FILL   = PatternFill("solid", fgColor="E8F0FE")
    TOTAL_FONT   = Font(name="Arial", bold=True, size=9, color="1E3A5F")
    BORDER_SIDE  = Side(style="thin", color="CCCCCC")
    THIN_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                          top=BORDER_SIDE, bottom=BORDER_SIDE)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT        = Alignment(horizontal="right",  vertical="center")

    # ── 1. Лист «Сводка» ──────────────────────────────
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_view.showGridLines = False

    # Заголовок
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Сформировано: {date.today().strftime('%d.%m.%Y')}  |  Всего документов: {len(documents)}"
    ws["A2"].font = Font(name="Arial", size=9, color="888888")
    ws["A2"].alignment = CENTER
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    # Заголовки таблицы
    headers = ["№", "Тип", "Номер", "Дата", "Поставщик GLN", "Покупатель GLN",
               "Позиций", "Сумма с НДС"]
    widths  = [5, 9, 18, 12, 16, 16, 9, 16]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 20

    # Данные
    total_amount = 0.0
    for i, doc in enumerate(documents, 1):
        row = i + 4
        raw = doc.raw_json or {}
        positions = raw.get("positions", [])
        amount = raw.get("totalAmount", 0) or 0

        values = [
            i,
            doc.doc_type,
            doc.number,
            doc.doc_date.strftime("%d.%m.%Y") if doc.doc_date else "—",
            doc.supplier_gln or "—",
            doc.buyer_gln or "—",
            len(positions),
            float(amount),
        ]
        fill = ALT_FILL if i % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font   = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = RIGHT if col in (1, 7, 8) else LEFT
            if fill:
                cell.fill = fill
            if col == 8:
                cell.number_format = '#,##0.00'
        total_amount += float(amount)

    # Итого
    total_row = len(documents) + 5
    ws.merge_cells(f"A{total_row}:G{total_row}")
    ws[f"A{total_row}"] = "ИТОГО"
    ws[f"A{total_row}"].font = TOTAL_FONT
    ws[f"A{total_row}"].fill = TOTAL_FILL
    ws[f"A{total_row}"].alignment = RIGHT
    ws[f"A{total_row}"].border = THIN_BORDER
    cell = ws.cell(row=total_row, column=8, value=total_amount)
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER
    cell.alignment = RIGHT
    cell.number_format = '#,##0.00'

    ws.freeze_panes = "A5"

    # ── 2. Листы по типам документов ─────────────────
    DOC_TYPE_NAMES = {
        "ORDER":  "Заказы",
        "ORDRSP": "Подтверждения",
        "DESADV": "Отгрузки",
        "INVOICE":"Счета-фактуры",
        "PRICAT": "Прайс-листы",
    }

    # Группируем по типу
    by_type = {}
    for doc in documents:
        by_type.setdefault(doc.doc_type, []).append(doc)

    for doc_type, docs in by_type.items():
        sheet_name = DOC_TYPE_NAMES.get(doc_type, doc_type)[:31]
        ws2 = wb.create_sheet(title=sheet_name)
        ws2.sheet_view.showGridLines = False

        # Заголовок листа
        ws2.merge_cells("A1:J1")
        ws2["A1"] = f"{sheet_name} — {len(docs)} шт."
        ws2["A1"].font = TITLE_FONT
        ws2["A1"].alignment = CENTER
        ws2.row_dimensions[1].height = 26

        # Шапка
        hdrs2 = ["№", "Номер", "Дата", "Поставщик GLN",
                 "Покупатель GLN", "EAN товара", "Наименование",
                 "Кол-во", "Цена", "Сумма с НДС"]
        ws2_widths = [5, 16, 12, 16, 16, 16, 30, 10, 12, 14]
        for col, (h, w) in enumerate(zip(hdrs2, ws2_widths), 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = THIN_BORDER
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.row_dimensions[3].height = 20

        row_num = 4
        for idx, doc in enumerate(docs, 1):
            raw = doc.raw_json or {}
            positions = raw.get("positions", []) or []

            if not positions:
                # Документ без позиций — одна строка
                values = [
                    idx, doc.number,
                    doc.doc_date.strftime("%d.%m.%Y") if doc.doc_date else "—",
                    doc.supplier_gln or "—",
                    doc.buyer_gln or "—",
                    "—", "—", "—", "—", raw.get("totalAmount", 0) or 0,
                ]
                fill = ALT_FILL if idx % 2 == 0 else None
                for col, val in enumerate(values, 1):
                    cell = ws2.cell(row=row_num, column=col, value=val)
                    cell.font = DATA_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = RIGHT if col in (1, 8, 9, 10) else LEFT
                    if fill:
                        cell.fill = fill
                    if col == 10:
                        cell.number_format = '#,##0.00'
                row_num += 1
            else:
                # Каждая позиция — отдельная строка
                fill = ALT_FILL if idx % 2 == 0 else None
                for p_idx, pos in enumerate(positions):
                    values = [
                        idx if p_idx == 0 else "",
                        doc.number if p_idx == 0 else "",
                        doc.doc_date.strftime("%d.%m.%Y") if (doc.doc_date and p_idx == 0) else ("—" if p_idx == 0 else ""),
                        doc.supplier_gln if p_idx == 0 else "",
                        doc.buyer_gln if p_idx == 0 else "",
                        str(pos.get("ean", "")),
                        str(pos.get("name", "")),
                        float(pos.get("qty", 0) or 0),
                        float(pos.get("price", 0) or 0),
                        float(pos.get("amount", 0) or 0),
                    ]
                    for col, val in enumerate(values, 1):
                        cell = ws2.cell(row=row_num, column=col, value=val)
                        cell.font = DATA_FONT
                        cell.border = THIN_BORDER
                        cell.alignment = RIGHT if col in (1, 8, 9, 10) else LEFT
                        if fill:
                            cell.fill = fill
                        if col in (9, 10):
                            cell.number_format = '#,##0.00'
                    row_num += 1

        ws2.freeze_panes = "A4"

    # ── Отдача файла ──────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"docrobot_export_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ═══════════════════════════════════════════════════════
# PDF
# ═══════════════════════════════════════════════════════

def export_pdf(documents, title="Печатная форма"):
    """Генерирует PDF — по одной странице на каждый документ."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, PageBreak, HRFlowable,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
    except ImportError:
        return HttpResponse("reportlab не установлен. Выполните: pip install reportlab", status=500)

    buf = io.BytesIO()

    # Шрифт с поддержкой кириллицы
    font_name = "Helvetica"
    try:
        font_paths = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                pdfmetrics.registerFont(TTFont("CyrFont", fp))
                font_name = "CyrFont"
                break
    except Exception:
        pass

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    # Стили
    W = A4[0] - 30*mm

    def sty(size=9, bold=False, color=colors.black, align="LEFT"):
        return ParagraphStyle(
            "s", fontName=font_name + ("-Bold" if bold and font_name != "Helvetica" else ""),
            fontSize=size, textColor=color,
            alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}.get(align, 0),
            leading=size * 1.3,
        )

    ACCENT = colors.HexColor("#1E3A5F")
    LIGHT  = colors.HexColor("#EFF4FB")

    story = []

    for i, doc in enumerate(documents):
        if i > 0:
            story.append(PageBreak())

        raw = doc.raw_json or {}
        positions = raw.get("positions", []) or []

        # ── Шапка документа ──
        doc_type_ru = {
            "ORDER": "ЗАКАЗ",
            "ORDRSP": "ПОДТВЕРЖДЕНИЕ ЗАКАЗА",
            "DESADV": "УВЕДОМЛЕНИЕ ОБ ОТГРУЗКЕ",
            "INVOICE": "СЧЁТ-ФАКТУРА",
            "PRICAT": "ПРАЙС-ЛИСТ",
        }.get(doc.doc_type, doc.doc_type)

        story.append(Paragraph(doc_type_ru, sty(16, bold=True, color=ACCENT, align="CENTER")))
        story.append(Spacer(1, 4*mm))

        # Реквизиты
        doc_date_str = doc.doc_date.strftime("%d.%m.%Y") if doc.doc_date else "—"
        info_data = [
            [Paragraph("Номер:", sty(9, bold=True, color=colors.HexColor("#666666"))),
             Paragraph(doc.number or "—", sty(10, bold=True)),
             Paragraph("Дата:", sty(9, bold=True, color=colors.HexColor("#666666"))),
             Paragraph(doc_date_str, sty(10, bold=True))],
            [Paragraph("Поставщик GLN:", sty(9, color=colors.HexColor("#666666"))),
             Paragraph(doc.supplier_gln or "—", sty(9)),
             Paragraph("Покупатель GLN:", sty(9, color=colors.HexColor("#666666"))),
             Paragraph(doc.buyer_gln or "—", sty(9))],
            [Paragraph("Поставщик:", sty(9, color=colors.HexColor("#666666"))),
             Paragraph(doc.supplier_name or "—", sty(9)),
             Paragraph("Покупатель:", sty(9, color=colors.HexColor("#666666"))),
             Paragraph(doc.buyer_name or "—", sty(9))],
        ]
        col_w = [W*0.2, W*0.3, W*0.2, W*0.3]
        info_tbl = Table(info_data, colWidths=col_w)
        info_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), LIGHT),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, colors.HexColor("#FAFCFF")]),
            ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(info_tbl)
        story.append(Spacer(1, 5*mm))

        # ── Таблица позиций ──
        if positions:
            tbl_headers = ["№", "EAN / Код", "Наименование товара", "Кол-во", "Ед.", "Цена", "Сумма с НДС"]
            col_w2 = [8*mm, 28*mm, W - 8-28-18-12-20-25*mm, 18*mm, 12*mm, 20*mm, 25*mm]

            tbl_data = [[Paragraph(h, sty(8, bold=True, color=colors.white, align="CENTER"))
                         for h in tbl_headers]]

            total = 0.0
            for j, pos in enumerate(positions, 1):
                amount = float(pos.get("amount", 0) or 0)
                total += amount
                tbl_data.append([
                    Paragraph(str(j), sty(8, align="CENTER")),
                    Paragraph(str(pos.get("ean", "")), sty(8)),
                    Paragraph(str(pos.get("name", ""))[:80], sty(8)),
                    Paragraph(str(pos.get("qty", 0)), sty(8, align="RIGHT")),
                    Paragraph(str(pos.get("unit", "шт")), sty(8, align="CENTER")),
                    Paragraph(f'{float(pos.get("price",0) or 0):,.2f}', sty(8, align="RIGHT")),
                    Paragraph(f'{amount:,.2f}', sty(8, align="RIGHT")),
                ])

            # Итого
            tbl_data.append([
                Paragraph("", sty(9)),
                Paragraph("", sty(9)),
                Paragraph("ИТОГО:", sty(9, bold=True, align="RIGHT")),
                Paragraph("", sty(9)),
                Paragraph("", sty(9)),
                Paragraph("", sty(9)),
                Paragraph(f'{total:,.2f}', sty(9, bold=True, align="RIGHT")),
            ])

            tbl = Table(tbl_data, colWidths=col_w2, repeatRows=1)
            tbl.setStyle(TableStyle([
                # Шапка
                ("BACKGROUND", (0,0), (-1,0), ACCENT),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("ALIGN", (0,0), (-1,0), "CENTER"),
                # Данные
                ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, colors.HexColor("#F5F8FF")]),
                # Итого
                ("BACKGROUND", (0,-1), (-1,-1), LIGHT),
                ("FONTNAME", (0,-1), (-1,-1), font_name),
                # Общее
                ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING", (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                ("LEFTPADDING", (0,0), (-1,-1), 4),
                ("RIGHTPADDING", (0,0), (-1,-1), 4),
            ]))
            story.append(tbl)
        else:
            story.append(Paragraph("Позиции товаров отсутствуют в сохранённых данных.", sty(9, color=colors.grey)))

        # Подпись
        story.append(Spacer(1, 8*mm))
        story.append(HRFlowable(width=W, thickness=0.5, color=colors.HexColor("#CCCCCC")))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(
            f"Сформировано системой DOCROBOT EDI Gateway · {date.today().strftime('%d.%m.%Y')}",
            sty(8, color=colors.HexColor("#AAAAAA"), align="CENTER")
        ))

    doc.build(story)
    buf.seek(0)

    filename = f"docrobot_{date.today().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ═══════════════════════════════════════════════════════
# XML
# ═══════════════════════════════════════════════════════

def export_xml_bundle(documents):
    """Генерирует XML-архив (zip) если несколько документов, иначе один XML."""
    from edi.xml_builder import build_xml

    if len(documents) == 1:
        doc = documents[0]
        try:
            xml_bytes = doc.xml_content.encode("utf-8") if doc.xml_content else build_xml(doc.doc_type, doc.raw_json)
        except Exception:
            xml_bytes = b"<?xml version='1.0' encoding='utf-8'?><error>XML generation failed</error>"
        filename = f"{doc.doc_type}_{doc.number or doc.docrobot_id}_{date.today().strftime('%Y%m%d')}.xml"
        filename = filename.replace("/", "-").replace(" ", "_")
        response = HttpResponse(xml_bytes, content_type="application/xml; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # Несколько — zip-архив
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for doc in documents:
            try:
                xml_bytes = doc.xml_content.encode("utf-8") if doc.xml_content else build_xml(doc.doc_type, doc.raw_json)
            except Exception:
                xml_bytes = b"<?xml version='1.0' encoding='utf-8'?><error>XML generation failed</error>"
            fname = f"{doc.doc_type}_{doc.number or doc.docrobot_id}.xml"
            fname = fname.replace("/", "-").replace(" ", "_")
            zf.writestr(fname, xml_bytes)
    buf.seek(0)
    filename = f"docrobot_xml_{date.today().strftime('%Y%m%d')}.zip"
    response = HttpResponse(buf.read(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
