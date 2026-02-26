"""
edi/views.py — Представления: дашборд, отчёты, API-действия

GET  /                   — дашборд
GET  /documents/         — список документов
GET  /documents/<id>/    — детали + XML-просмотр
GET  /queue/             — очередь отправки
GET  /logs/              — логи активности
GET  /reports/           — отчёты и статистика
GET  /reports/export/    — экспорт в CSV

POST /api/retry/<id>/    — повторить отправку вручную
POST /api/send/<id>/     — отправить документ вручную
POST /api/webhook/       — вебхук: принять документ от Docrobot
"""
import json
import csv
import io
import logging
from datetime import date, timedelta, datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response

from .models import EdiDocument, SendQueue, ActivityLog, DocumentComment
from .services import process_document, DocrobotClient
from .xml_builder import build_xml

logger = logging.getLogger('edi')


# ─── Дашборд ─────────────────────────────────────────────

def dashboard(request):
    # Статистика по статусам очереди
    queue_stats = {
        row['status']: row['cnt']
        for row in SendQueue.objects.values('status').annotate(cnt=Count('id'))
    }
    # Статистика по типам документов
    doc_stats = {
        row['doc_type']: row['cnt']
        for row in EdiDocument.objects.values('doc_type').annotate(cnt=Count('id'))
    }
    # Последние 8 документов
    recent_docs  = EdiDocument.objects.select_related('queue_entry').order_by('-received_at')[:8]
    # Ошибочные записи
    errors       = SendQueue.objects.filter(
        status__in=[SendQueue.STATUS_ERROR, SendQueue.STATUS_FAILED]
    ).select_related('document').order_by('-updated_at')[:5]
    # Последние 10 логов
    recent_logs  = ActivityLog.objects.order_by('-created_at')[:10]
    # Динамика за 7 дней
    seven_days = []
    for i in range(6, -1, -1):
        day = date.today() - timedelta(days=i)
        cnt = EdiDocument.objects.filter(received_at__date=day).count()
        seven_days.append({'day': day.strftime('%d.%m'), 'count': cnt})

    return render(request, 'edi/dashboard.html', {
        'queue_stats': queue_stats,
        'doc_stats':   doc_stats,
        'recent_docs': recent_docs,
        'errors':      errors,
        'recent_logs': recent_logs,
        'seven_days':  json.dumps(seven_days),
        'total_docs':  EdiDocument.objects.count(),
        'total_sent':  SendQueue.objects.filter(status=SendQueue.STATUS_SENT).count(),
        'total_errors':SendQueue.objects.filter(status__in=[SendQueue.STATUS_ERROR, SendQueue.STATUS_FAILED]).count(),
        'pending':     SendQueue.objects.filter(status=SendQueue.STATUS_PENDING).count(),
    })


# ─── Документы ───────────────────────────────────────────

def documents(request):
    from django.core.paginator import Paginator
    qs = EdiDocument.objects.select_related('queue_entry').order_by('-received_at')

    doc_type  = request.GET.get('type', '')
    search    = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    if doc_type:
        qs = qs.filter(doc_type=doc_type)
    if search:
        qs = qs.filter(
            Q(number__icontains=search) |
            Q(supplier_name__icontains=search) |
            Q(buyer_name__icontains=search) |
            Q(supplier_gln__icontains=search) |
            Q(buyer_gln__icontains=search)
        )
    if date_from:
        qs = qs.filter(received_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(received_at__date__lte=date_to)

    paginator = Paginator(qs, 50)
    page_num  = request.GET.get('page', 1)
    page_obj  = paginator.get_page(page_num)

    return render(request, 'edi/documents.html', {
        'documents': page_obj,
        'page_obj':  page_obj,
        'doc_types': EdiDocument.DOC_TYPES,
        'filters': {'type': doc_type, 'q': search, 'date_from': date_from, 'date_to': date_to},
        'total': paginator.count,
    })


def document_detail(request, pk):
    doc   = get_object_or_404(EdiDocument, pk=pk)
    queue = getattr(doc, 'queue_entry', None)
    logs  = doc.logs.order_by('-created_at')
    comments = doc.comments.order_by('-created_at')

    # Генерируем XML для просмотра если нет
    if not doc.xml_content and doc.raw_json:
        try:
            doc.xml_content = build_xml(doc.doc_type, doc.raw_json).decode('utf-8')
            doc.save(update_fields=['xml_content'])
        except Exception:
            pass

    # Передаём позиции явно — Django шаблон иногда не читает вложенные JSONField ключи
    raw = doc.raw_json or {}
    positions = raw.get('positions') or []

    return render(request, 'edi/document_detail.html', {
        'doc': doc, 'queue': queue, 'logs': logs, 'comments': comments,
        'importance_choices': DocumentComment.IMPORTANCE_CHOICES,
        'positions': positions,
        'total_amount': raw.get('totalAmount', ''),
    })


def api_comment_add(request, pk):
    """POST /api/comments/<pk>/add/ — добавить комментарий к документу."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    doc  = get_object_or_404(EdiDocument, pk=pk)
    text = request.POST.get('text', '').strip()
    if not text:
        return JsonResponse({'ok': False, 'error': 'Текст не может быть пустым'}, status=400)

    importance = request.POST.get('importance', DocumentComment.IMPORTANCE_NORMAL)
    author     = request.POST.get('author', '').strip() or 'Оператор'

    comment = DocumentComment.objects.create(
        document=doc, text=text, importance=importance, author=author
    )
    logger.info(f'Комментарий к {doc.doc_type} №{doc.number}: [{importance}] {text[:60]}')
    return JsonResponse({
        'ok': True,
        'id':         comment.pk,
        'text':       comment.text,
        'importance': comment.importance,
        'author':     comment.author,
        'created_at': comment.created_at.strftime('%d.%m.%Y %H:%M'),
    })


def api_comment_delete(request, comment_id):
    """POST /api/comments/<id>/delete/ — удалить комментарий."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    comment = get_object_or_404(DocumentComment, pk=comment_id)
    comment.delete()
    return JsonResponse({'ok': True})


# ─── Очередь ─────────────────────────────────────────────

def queue(request):
    qs = SendQueue.objects.select_related('document').order_by('-updated_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        qs = qs.filter(status=status_filter)
    return render(request, 'edi/queue.html', {
        'entries': qs[:100],
        'statuses': SendQueue.STATUSES,
        'filter_status': status_filter,
    })


# ─── Логи ────────────────────────────────────────────────

def logs(request):
    from django.core.paginator import Paginator
    qs = ActivityLog.objects.select_related('document').order_by('-created_at')
    level  = request.GET.get('level', '')
    action = request.GET.get('action', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    if level:
        qs = qs.filter(level=level)
    if action:
        qs = qs.filter(action__icontains=action)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # Экспорт в Excel
    if request.GET.get('export') == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl не установлен', status=500)

        import io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Логи'
        ws.sheet_view.showGridLines = False

        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Логи активности Docrobot EDI — {date.today().strftime("%d.%m.%Y")}'
        ws['A1'].font = Font(name='Arial', bold=True, size=13, color='1E3A5F')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 24

        # Шапка
        HFILL = PatternFill('solid', fgColor='1E3A5F')
        HFONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        headers = ['Время', 'Уровень', 'Действие', 'Сообщение', 'Документ', 'ID документа']
        widths  = [20, 10, 20, 60, 20, 20]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = HFONT; cell.fill = HFILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[3].height = 18

        ALT = PatternFill('solid', fgColor='F5F8FF')
        DFONT = Font(name='Arial', size=9)
        LEVEL_COLORS = {'error': 'FEE2E2', 'warn': 'FEF3C7', 'info': 'F0FDF4'}

        for i, log in enumerate(qs[:5000], 1):
            row = i + 3
            fill = PatternFill('solid', fgColor=LEVEL_COLORS.get(log.level, 'FFFFFF')) if log.level in LEVEL_COLORS else (ALT if i % 2 == 0 else None)
            vals = [
                log.created_at.strftime('%d.%m.%Y %H:%M:%S'),
                log.level.upper(),
                log.action,
                log.message,
                log.document.number if log.document else '—',
                log.document.docrobot_id if log.document else '—',
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = DFONT
                if fill: cell.fill = fill
                cell.alignment = Alignment(vertical='center', wrap_text=(col == 4))
            ws.row_dimensions[row].height = 15

        ws.freeze_panes = 'A4'
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f'logs_{date.today().strftime("%Y%m%d")}.xlsx'
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    paginator = Paginator(qs, 100)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'edi/logs.html', {
        'logs': page_obj,
        'page_obj': page_obj,
        'total': paginator.count,
        'levels': ActivityLog.LEVELS,
        'filter_level': level,
        'filter_action': action,
        'filter_date_from': date_from,
        'filter_date_to':   date_to,
    })


# ─── Отчёты ──────────────────────────────────────────────

def reports(request):
    # Период
    days = int(request.GET.get('days', 30))
    since = timezone.now() - timedelta(days=days)

    # Документы по типам за период
    by_type = list(
        EdiDocument.objects.filter(received_at__gte=since)
        .values('doc_type').annotate(cnt=Count('id'))
        .order_by('-cnt')
    )
    # Успешно / с ошибкой
    sent_count   = SendQueue.objects.filter(status=SendQueue.STATUS_SENT, updated_at__gte=since).count()
    failed_count = SendQueue.objects.filter(status=SendQueue.STATUS_FAILED, updated_at__gte=since).count()
    error_count  = SendQueue.objects.filter(status=SendQueue.STATUS_ERROR, updated_at__gte=since).count()

    # Топ поставщиков
    top_suppliers = list(
        EdiDocument.objects.filter(received_at__gte=since)
        .values('supplier_name').annotate(cnt=Count('id'))
        .order_by('-cnt')[:10]
    )
    # Динамика по дням
    daily = []
    for i in range(days - 1, -1, -1):
        day = date.today() - timedelta(days=i)
        cnt = EdiDocument.objects.filter(received_at__date=day).count()
        err = SendQueue.objects.filter(updated_at__date=day, status__in=[SendQueue.STATUS_ERROR, SendQueue.STATUS_FAILED]).count()
        daily.append({'day': day.strftime('%d.%m'), 'count': cnt, 'errors': err})

    return render(request, 'edi/reports.html', {
        'days': days,
        'by_type': by_type,
        'sent_count': sent_count,
        'failed_count': failed_count,
        'error_count': error_count,
        'top_suppliers': top_suppliers,
        'daily_json': json.dumps(daily),
        'total_period': EdiDocument.objects.filter(received_at__gte=since).count(),
    })


def reports_export(request):
    """Экспорт всех документов за период в CSV (UTF-8 с BOM для Excel)."""
    days  = int(request.GET.get('days', 30))
    since = timezone.now() - timedelta(days=days)
    docs  = EdiDocument.objects.filter(received_at__gte=since).select_related('queue_entry').order_by('-received_at')

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['ID Docrobot', 'Тип', 'Номер', 'Дата', 'Поставщик', 'Покупатель', 'Статус очереди', 'Попыток', 'Получен'])
    for d in docs:
        q = getattr(d, 'queue_entry', None)
        writer.writerow([
            d.docrobot_id, d.get_doc_type_display(), d.number,
            d.doc_date or '', d.supplier_name, d.buyer_name,
            q.get_status_display() if q else '—',
            q.attempts if q else 0,
            d.received_at.strftime('%d.%m.%Y %H:%M'),
        ])

    content = '\ufeff' + buf.getvalue()  # BOM для Excel
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="edi-report-{date.today()}.csv"'
    return response


# ─── API: ручные действия ────────────────────────────────

@api_view(['POST'])
def api_retry(request, pk):
    """Повторить отправку записи очереди вручную."""
    entry = get_object_or_404(SendQueue, pk=pk)
    if entry.status in [SendQueue.STATUS_SENT]:
        return Response({'error': 'Документ уже отправлен'}, status=400)

    entry.status     = SendQueue.STATUS_PENDING
    entry.next_retry = None
    entry.save(update_fields=['status', 'next_retry'])

    success = process_document(entry)
    entry.refresh_from_db()
    return Response({
        'success': success,
        'status':  entry.status,
        'error':   entry.last_error,
    })


@api_view(['POST'])
def api_send_document(request, pk):
    """Отправить конкретный документ в 1С."""
    doc   = get_object_or_404(EdiDocument, pk=pk)
    entry, _ = SendQueue.objects.get_or_create(document=doc)
    entry.status     = SendQueue.STATUS_PENDING
    entry.next_retry = None
    entry.save(update_fields=['status', 'next_retry'])
    success = process_document(entry)
    entry.refresh_from_db()
    return Response({'success': success, 'status': entry.status, 'error': entry.last_error})


@csrf_exempt
@require_POST
def api_webhook(request):
    """
    Вебхук: Docrobot отправляет документ → сохраняем → добавляем в очередь.
    POST /api/webhook/
    Body: JSON с полями документа
    """
    try:
        raw = json.loads(request.body)
        client = DocrobotClient()
        normalized = client.normalize_document(raw)
        doc_id = normalized['docrobotId']

        if not doc_id:
            return JsonResponse({'error': 'docrobotId обязателен'}, status=400)

        doc, created = EdiDocument.objects.get_or_create(
            docrobot_id=doc_id,
            defaults={
                'doc_type':      normalized['docType'],
                'number':        normalized['number'],
                'doc_date':      normalized['date'] or None,
                'supplier_gln':  normalized['supplierGln'],
                'buyer_gln':     normalized['buyerGln'],
                'supplier_name': normalized['supplierName'],
                'buyer_name':    normalized['buyerName'],
                'raw_json':      normalized['raw'],
            }
        )
        if created:
            SendQueue.objects.create(document=doc)
            ActivityLog.objects.create(
                level='info', action='webhook_received',
                message=f'Получен документ {doc}', document=doc,
            )
        return JsonResponse({'saved': created, 'id': doc.id})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ─── Настройки: XML-шаблоны ──────────────────────────────

def settings_view(request):
    """Страница настроек: шаблоны XML, параметры 1С."""
    from .models import XmlTemplate
    from .xml_builder import DEFAULT_TEMPLATES

    # Авто-создание дефолтных шаблонов при первом открытии
    for doc_type, tpl_data in DEFAULT_TEMPLATES.items():
        XmlTemplate.objects.get_or_create(
            doc_type=doc_type,
            defaults={
                'name':         tpl_data['name'],
                'position_tpl': tpl_data['position_tpl'],
                'body_tpl':     tpl_data['body_tpl'],
            }
        )

    templates = XmlTemplate.objects.all().order_by('doc_type')

    VARS = [
        ('{{number}}',         'Номер документа'),
        ('{{date}}',           'Дата документа'),
        ('{{delivery_date}}',  'Дата доставки'),
        ('{{supplier_gln}}',   'GLN поставщика'),
        ('{{supplier_name}}',  'Название поставщика'),
        ('{{buyer_gln}}',      'GLN покупателя'),
        ('{{buyer_name}}',     'Название покупателя'),
        ('{{currency}}',       'Валюта'),
        ('{{order_number}}',   'Номер заказа-основания'),
        ('{{shipment_date}}',  'Дата отгрузки'),
        ('{{total_amount}}',   'Сумма без НДС'),
        ('{{total_vat}}',      'Сумма НДС'),
        ('{{total_with_vat}}', 'Сумма с НДС'),
        ('{{positions}}',      'Блок позиций (из шаблона позиции)'),
        ('{{positions_json}}', 'Позиции как JSON'),
        ('{{raw_json}}',       'Весь документ как JSON'),
    ]

    return render(request, 'edi/settings.html', {'templates': templates, 'vars': VARS})


def template_edit(request, doc_type):
    """Редактирование конкретного XML-шаблона."""
    from .models import XmlTemplate
    tpl = get_object_or_404(XmlTemplate, doc_type=doc_type)

    if request.method == 'POST':
        tpl.name         = request.POST.get('name', tpl.name)
        tpl.body_tpl     = request.POST.get('body_tpl', tpl.body_tpl)
        tpl.position_tpl = request.POST.get('position_tpl', tpl.position_tpl)
        tpl.content_type = request.POST.get('content_type', tpl.content_type)
        tpl.is_active    = request.POST.get('is_active') == 'on'
        tpl.save()
        ActivityLog.objects.create(
            level='info', action='template_updated',
            message=f'Шаблон {doc_type} обновлён',
        )
        return redirect('settings')

    doc_vars = [
        'number', 'date', 'delivery_date', 'supplier_gln', 'supplier_name',
        'buyer_gln', 'buyer_name', 'currency', 'order_number', 'shipment_date',
        'total_amount', 'total_vat', 'total_with_vat', 'positions', 'positions_json', 'raw_json',
    ]
    pos_vars = ['line', 'ean', 'item_code', 'item_name', 'quantity', 'unit_price', 'vat', 'amount', 'amount_with_vat']

    return render(request, 'edi/template_edit.html', {
        'tpl': tpl, 'doc_vars': doc_vars, 'pos_vars': pos_vars,
    })


@api_view(['POST'])
def api_test_xml(request):
    """
    Рендерит шаблон с тестовыми данными и возвращает XML.
    Тело запроса: { doc_type, body_tpl, position_tpl, sample_data (опц.) }
    """
    from .models import XmlTemplate

    doc_type     = request.data.get('doc_type', 'ORDER')
    body_tpl     = request.data.get('body_tpl', '')
    position_tpl = request.data.get('position_tpl', '')

    sample = request.data.get('sample_data') or {
        'number':       'TEST-001',
        'date':         '2026-02-20',
        'deliveryDate': '2026-02-25',
        'supplierGln':  '4600000000001',
        'supplierName': 'ТОО Поставщик',
        'buyerGln':     '4600000000002',
        'buyerName':    'ТОО Покупатель',
        'currency':     'KZT',
        'orderNumber':  'ORD-001',
        'totalAmount':  '10000',
        'totalVat':     '1200',
        'totalWithVat': '11200',
        'positions': [
            {
                'ean': '4600123456789', 'itemCode': 'SKU-001',
                'itemName': 'Товар тестовый', 'quantity': 10,
                'unitPrice': 1000, 'vat': 12, 'amount': 10000, 'amountWithVat': 11200,
            }
        ],
    }

    tpl = XmlTemplate(doc_type=doc_type, name='test', body_tpl=body_tpl, position_tpl=position_tpl)
    try:
        xml = tpl.render(sample)
        return Response({'xml': xml, 'success': True})
    except Exception as e:
        return Response({'error': str(e), 'success': False}, status=400)


@api_view(['POST'])
def api_test_send(request):
    """
    Отправляет тестовый XML в 1С и возвращает результат.
    Тело: { doc_type, xml (строка) }
    """
    from .services import OneCClient
    doc_type = request.data.get('doc_type', 'ORDER')
    xml_str  = request.data.get('xml', '')
    if not xml_str:
        return Response({'error': 'xml обязателен'}, status=400)
    try:
        client = OneCClient()
        code, resp = client.send(xml_str.encode('utf-8'), doc_type)
        ActivityLog.objects.create(
            level='info' if 200 <= code < 300 else 'warn',
            action='test_send',
            message=f'Тест {doc_type}: HTTP {code} → {resp[:200]}',
        )
        return Response({'http_status': code, 'response': resp, 'success': 200 <= code < 300})
    except Exception as e:
        return Response({'error': str(e), 'success': False}, status=500)


# ─── Страница подключений ─────────────────────────────────

def connections_view(request):
    """Страница настройки подключений к Docrobot и 1С."""
    from .models import ConnectionSettings
    cfg = ConnectionSettings.get()

    if request.method == 'POST':
        cfg.docrobot_url           = request.POST.get('docrobot_url', cfg.docrobot_url).strip()
        cfg.docrobot_username      = request.POST.get('docrobot_username', '').strip()
        new_dr_pass                = request.POST.get('docrobot_password', '').strip()
        if new_dr_pass:
            cfg.docrobot_password  = new_dr_pass
        cfg.docrobot_poll_interval = int(request.POST.get('docrobot_poll_interval', 60))
        cfg.docrobot_gln           = request.POST.get('docrobot_gln', '').strip()
        cfg.cleanup_days           = int(request.POST.get('cleanup_days', 90) or 90)

        cfg.onec_url      = request.POST.get('onec_url', cfg.onec_url).strip()
        cfg.onec_username = request.POST.get('onec_username', '').strip()
        new_1c_pass       = request.POST.get('onec_password', '').strip()
        if new_1c_pass:
            cfg.onec_password = new_1c_pass
        cfg.onec_timeout  = int(request.POST.get('onec_timeout', 30))

        cfg.telegram_token   = request.POST.get('telegram_token', '').strip()
        cfg.telegram_chat_id = request.POST.get('telegram_chat_id', '').strip()

        cfg.save()
        cfg.apply_to_django_settings()

        ActivityLog.objects.create(
            level='info', action='settings_saved',
            message='Настройки подключений обновлены',
        )
        return redirect('connections')

    return render(request, 'edi/connections.html', {'cfg': cfg})


@api_view(['POST'])
def api_test_docrobot(request):
    """Тест авторизации в Docrobot — использует данные из запроса или из БД."""
    from .models import ConnectionSettings
    import requests as req

    url      = request.data.get('url', '').strip()
    username = request.data.get('username', '').strip()
    password = request.data.get('password', '').strip()

    # Если пароль не передан — берём из БД
    if not password:
        cfg = ConnectionSettings.get()
        password = cfg.docrobot_password

    if not url or not username:
        return Response({'success': False, 'error': 'URL и логин обязательны'}, status=400)

    try:
        resp = req.post(
            f'{url.rstrip("/")}/api/v1/auth',
            json={'login': username, 'password': password},
            headers={'Content-type': 'application/json'},
            timeout=15,
        )
        data = resp.json() if resp.content else {}

        if resp.status_code == 200:
            token = data.get('token') or data.get('access_token') or data.get('accessToken', '')
            # Сохраняем статус
            from .models import ConnectionSettings
            from django.utils import timezone
            cfg = ConnectionSettings.get()
            cfg.docrobot_status    = 'ok'
            cfg.docrobot_tested_at = timezone.now()
            cfg.save(update_fields=['docrobot_status', 'docrobot_tested_at'])

            ActivityLog.objects.create(level='info', action='docrobot_auth_ok',
                message=f'Авторизация Docrobot успешна: {username}@{url}')
            return Response({'success': True, 'token_preview': token[:20] + '...' if token else '(пусто)', 'http_status': 200})
        else:
            cfg = ConnectionSettings.get()
            cfg.docrobot_status = 'error'
            cfg.save(update_fields=['docrobot_status'])
            ActivityLog.objects.create(level='error', action='docrobot_auth_fail',
                message=f'Ошибка авторизации Docrobot HTTP {resp.status_code}: {resp.text[:200]}')
            return Response({'success': False, 'error': f'HTTP {resp.status_code}: {resp.text[:300]}', 'http_status': resp.status_code})

    except Exception as e:
        return Response({'success': False, 'error': str(e)}, status=500)


@api_view(['POST'])
def api_test_onec(request):
    """Тест подключения к 1С HTTP-сервису."""
    import requests as req

    url      = request.data.get('url', '').strip()
    username = request.data.get('username', '').strip()
    password = request.data.get('password', '').strip()

    if not password:
        from .models import ConnectionSettings
        cfg = ConnectionSettings.get()
        password = cfg.onec_password

    if not url:
        return Response({'success': False, 'error': 'URL обязателен'}, status=400)

    try:
        auth = (username, password) if username else None
        resp = req.get(
            url,
            auth=auth,
            headers={'Accept': 'application/xml, text/xml, */*'},
            timeout=10,
        )
        from .models import ConnectionSettings
        from django.utils import timezone
        cfg = ConnectionSettings.get()

        # 1С может ответить 200, 400, 405 — всё это значит "доступен"
        reachable = resp.status_code < 500
        cfg.onec_status    = 'ok' if reachable else 'error'
        cfg.onec_tested_at = timezone.now()
        cfg.save(update_fields=['onec_status', 'onec_tested_at'])

        ActivityLog.objects.create(
            level='info' if reachable else 'error',
            action='onec_test',
            message=f'Тест 1С {url}: HTTP {resp.status_code}',
        )
        return Response({
            'success':     reachable,
            'http_status': resp.status_code,
            'response':    resp.text[:300],
            'note':        'Статус < 500 означает что сервер доступен' if reachable else 'Сервер недоступен',
        })
    except req.exceptions.ConnectionError:
        return Response({'success': False, 'error': 'Не удалось подключиться — проверьте URL и что 1С запущен'}, status=200)
    except req.exceptions.Timeout:
        return Response({'success': False, 'error': 'Таймаут — 1С не отвечает за 10 секунд'}, status=200)
    except Exception as e:
        return Response({'success': False, 'error': str(e)}, status=500)


# ═══════════════════════════════════════════════════════
# Печатные формы
# ═══════════════════════════════════════════════════════

def print_forms(request):
    """Страница печатных форм с фильтрами и экспортом."""
    from .export import export_xlsx, export_pdf, export_xml_bundle
    from django.db.models import Count

    qs = EdiDocument.objects.order_by('-received_at')

    # Фильтры
    doc_type  = request.GET.get('doc_type', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    if doc_type:
        qs = qs.filter(doc_type=doc_type)
    if date_from:
        qs = qs.filter(doc_date__gte=date_from)
    if date_to:
        qs = qs.filter(doc_date__lte=date_to)

    fmt = request.GET.get('export', '')
    if fmt in ('xlsx', 'pdf', 'xml'):
        docs = list(qs)
        if fmt == 'xlsx':
            title = f"Документы EDI — {date_from or '...'} — {date_to or '...'}"
            return export_xlsx(docs, title=title)
        elif fmt == 'pdf':
            return export_pdf(docs)
        elif fmt == 'xml':
            return export_xml_bundle(docs)

    documents = list(qs[:500])
    type_counts = {}
    for d in documents:
        type_counts[d.doc_type] = type_counts.get(d.doc_type, 0) + 1

    return render(request, 'edi/print_forms.html', {
        'documents':   documents,
        'type_counts': type_counts,
    })


def print_single(request, pk, fmt):
    """Скачать один документ в нужном формате."""
    from .export import export_xlsx, export_pdf, export_xml_bundle
    doc = get_object_or_404(EdiDocument, pk=pk)
    docs = [doc]
    if fmt == 'xlsx':
        return export_xlsx(docs, title=f"{doc.doc_type} №{doc.number}")
    elif fmt == 'pdf':
        return export_pdf(docs)
    elif fmt == 'xml':
        return export_xml_bundle(docs)
    return HttpResponse("Неизвестный формат", status=400)


def print_selected(request):
    """Экспорт выбранных документов (по списку ID)."""
    from .export import export_xlsx, export_pdf, export_xml_bundle
    ids = request.GET.getlist('ids')
    fmt = request.GET.get('export', 'xlsx')
    docs = list(EdiDocument.objects.filter(pk__in=ids).order_by('-received_at'))
    if not docs:
        return HttpResponse("Документы не найдены", status=404)
    if fmt == 'xlsx':
        return export_xlsx(docs, title=f"Выбранные документы ({len(docs)} шт.)")
    elif fmt == 'pdf':
        return export_pdf(docs)
    elif fmt == 'xml':
        return export_xml_bundle(docs)
    return HttpResponse("Неизвестный формат", status=400)


# ═══════════════════════════════════════════════════════
# API: Живой поиск
# ═══════════════════════════════════════════════════════

def api_search(request):
    """GET /api/search/?q=...&type=...&limit=20 — JSON для live-поиска."""
    q         = request.GET.get('q', '').strip()
    doc_type  = request.GET.get('type', '')
    limit     = min(int(request.GET.get('limit', 20)), 100)

    qs = EdiDocument.objects.select_related('queue_entry').order_by('-received_at')

    if doc_type:
        qs = qs.filter(doc_type=doc_type)
    if q:
        qs = qs.filter(
            Q(number__icontains=q) |
            Q(supplier_name__icontains=q) |
            Q(buyer_name__icontains=q) |
            Q(supplier_gln__icontains=q) |
            Q(buyer_gln__icontains=q) |
            Q(docrobot_id__icontains=q)
        )

    docs = qs[:limit]
    results = []
    for doc in docs:
        q_entry = getattr(doc, 'queue_entry', None)
        results.append({
            'id':           doc.pk,
            'doc_type':     doc.doc_type,
            'number':       doc.number or '—',
            'doc_date':     doc.doc_date.strftime('%d.%m.%Y') if doc.doc_date else '—',
            'supplier_gln': doc.supplier_gln or '—',
            'buyer_gln':    doc.buyer_gln or '—',
            'supplier_name':doc.supplier_name or '—',
            'status':       q_entry.status if q_entry else 'none',
            'status_label': q_entry.get_status_display() if q_entry else '—',
            'received_at':  doc.received_at.strftime('%d.%m.%Y %H:%M'),
            'url':          f'/documents/{doc.pk}/',
        })

    return JsonResponse({'results': results, 'total': qs.count(), 'shown': len(results)})


# ═══════════════════════════════════════════════════════
# API: Дашборд реального времени
# ═══════════════════════════════════════════════════════

def api_dashboard_stats(request):
    """GET /api/dashboard/stats/ — JSON со свежей статистикой для авто-обновления."""
    from django.db.models import Count
    from datetime import timedelta

    queue_stats = {
        row['status']: row['cnt']
        for row in SendQueue.objects.values('status').annotate(cnt=Count('id'))
    }
    doc_stats = {
        row['doc_type']: row['cnt']
        for row in EdiDocument.objects.values('doc_type').annotate(cnt=Count('id'))
    }

    # Последние 5 документов
    recent = []
    for doc in EdiDocument.objects.select_related('queue_entry').order_by('-received_at')[:5]:
        q = getattr(doc, 'queue_entry', None)
        recent.append({
            'id':         doc.pk,
            'doc_type':   doc.doc_type,
            'number':     doc.number or '—',
            'status':     q.status if q else 'none',
            'status_label': q.get_status_display() if q else '—',
            'received_at': doc.received_at.strftime('%d.%m.%Y %H:%M'),
        })

    # Последний лог
    last_logs = []
    for log in ActivityLog.objects.order_by('-created_at')[:5]:
        last_logs.append({
            'level':   log.level,
            'action':  log.action,
            'message': log.message[:100],
            'time':    log.created_at.strftime('%H:%M:%S'),
        })

    # Динамика за 7 дней
    seven_days = []
    for i in range(6, -1, -1):
        day = date.today() - timedelta(days=i)
        cnt = EdiDocument.objects.filter(received_at__date=day).count()
        seven_days.append({'day': day.strftime('%d.%m'), 'count': cnt})

    return JsonResponse({
        'queue_stats':  queue_stats,
        'doc_stats':    doc_stats,
        'total_docs':   EdiDocument.objects.count(),
        'total_sent':   SendQueue.objects.filter(status=SendQueue.STATUS_SENT).count(),
        'total_errors': SendQueue.objects.filter(status__in=[SendQueue.STATUS_ERROR, SendQueue.STATUS_FAILED]).count(),
        'pending':      SendQueue.objects.filter(status=SendQueue.STATUS_PENDING).count(),
        'recent_docs':  recent,
        'last_logs':    last_logs,
        'seven_days':   seven_days,
        'server_time':  timezone.now().strftime('%d.%m.%Y %H:%M:%S'),
    })


# ═══════════════════════════════════════════════════════
# Кнопка «Получить сейчас»
# ═══════════════════════════════════════════════════════

@require_POST
def api_poll_now(request):
    """POST /api/poll-now/ — немедленный запуск одного цикла поллинга."""
    import threading
    from .services import DocrobotClient
    from .models import EdiDocument, SendQueue, ActivityLog

    def _run():
        try:
            client = DocrobotClient()
            documents = client.get_incoming_documents()
            new_count = 0
            for normalized in documents:
                doc_id = normalized.get('docrobotId', '')
                if not doc_id:
                    continue
                if EdiDocument.objects.filter(docrobot_id=doc_id).exists():
                    continue
                doc = EdiDocument.objects.create(
                    docrobot_id   = doc_id,
                    doc_type      = normalized['docType'],
                    number        = normalized.get('number', ''),
                    doc_date      = normalized.get('date') or None,
                    supplier_gln  = normalized.get('supplierGln', ''),
                    buyer_gln     = normalized.get('buyerGln', ''),
                    supplier_name = normalized.get('supplierName', ''),
                    buyer_name    = normalized.get('buyerName', ''),
                    raw_json      = normalized,
                )
                SendQueue.objects.create(document=doc)
                new_count += 1
            ActivityLog.objects.create(
                level='info', action='manual_poll',
                message=f'Ручной поллинг: получено {new_count} новых документов',
            )
        except Exception as e:
            ActivityLog.objects.create(
                level='error', action='manual_poll',
                message=f'Ошибка ручного поллинга: {e}',
            )

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return JsonResponse({'ok': True, 'message': 'Поллинг запущен в фоне'})


# ═══════════════════════════════════════════════════════
# Healthcheck
# ═══════════════════════════════════════════════════════

def healthcheck(request):
    """GET /health/ — страница статуса всех компонентов системы."""
    import platform
    import sys
    from django.db import connection
    from .models import ConnectionSettings, ActivityLog

    checks = []

    # 1. База данных
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM edi_edidocument")
            doc_count = cur.fetchone()[0]
        checks.append({'name': 'База данных (SQLite)', 'status': 'ok',
                        'detail': f'{doc_count} документов', 'icon': '🗄️'})
    except Exception as e:
        checks.append({'name': 'База данных (SQLite)', 'status': 'error',
                        'detail': str(e), 'icon': '🗄️'})

    # 2. Docrobot API — ping auth endpoint
    try:
        import requests as req
        cfg = ConnectionSettings.get()
        r = req.post(
            'https://edi-api.docrobot.kz/api/v1/auth',
            json={'login': cfg.docrobot_username, 'password': cfg.docrobot_password},
            timeout=8,
        )
        if r.status_code == 200 and r.json().get('checkStatus') == 0:
            checks.append({'name': 'Docrobot API', 'status': 'ok',
                            'detail': f'Авторизация успешна · {cfg.docrobot_username}', 'icon': '🔗'})
        else:
            checks.append({'name': 'Docrobot API', 'status': 'warn',
                            'detail': f'HTTP {r.status_code} · checkStatus={r.json().get("checkStatus")}', 'icon': '🔗'})
    except Exception as e:
        checks.append({'name': 'Docrobot API', 'status': 'error',
                        'detail': str(e)[:120], 'icon': '🔗'})

    # 3. 1С HTTP-сервис
    try:
        import requests as req
        cfg = ConnectionSettings.get()
        if cfg.onec_url and cfg.onec_url != 'http://localhost/hs/docrobot/orders':
            auth = (cfg.onec_username, cfg.onec_password) if cfg.onec_username else None
            r = req.get(cfg.onec_url, auth=auth, timeout=5)
            # 404/405 — сервис есть, но неверный метод/путь — это нормально
            if r.status_code < 500:
                checks.append({'name': '1С HTTP-сервис', 'status': 'ok',
                                'detail': f'HTTP {r.status_code} · {cfg.onec_url}', 'icon': '1️⃣'})
            else:
                checks.append({'name': '1С HTTP-сервис', 'status': 'error',
                                'detail': f'HTTP {r.status_code} · сервер вернул ошибку', 'icon': '1️⃣'})
        else:
            checks.append({'name': '1С HTTP-сервис', 'status': 'warn',
                            'detail': 'URL не настроен — перейдите в Подключения', 'icon': '1️⃣'})
    except Exception as e:
        checks.append({'name': '1С HTTP-сервис', 'status': 'error',
                        'detail': str(e)[:120], 'icon': '1️⃣'})

    # 4. Последний поллинг
    try:
        last_poll = ActivityLog.objects.filter(
            action__in=['docrobot_poll', 'manual_poll']
        ).order_by('-created_at').first()
        if last_poll:
            delta = timezone.now() - last_poll.created_at
            mins  = int(delta.total_seconds() // 60)
            status = 'ok' if mins < 10 else ('warn' if mins < 60 else 'error')
            checks.append({'name': 'Последний поллинг', 'status': status,
                            'detail': f'{mins} мин назад · {last_poll.message[:80]}', 'icon': '🔄'})
        else:
            checks.append({'name': 'Последний поллинг', 'status': 'warn',
                            'detail': 'Поллинг ещё не запускался', 'icon': '🔄'})
    except Exception as e:
        checks.append({'name': 'Последний поллинг', 'status': 'error',
                        'detail': str(e), 'icon': '🔄'})

    # 5. Очередь — зависшие документы
    try:
        from datetime import timedelta
        stuck = SendQueue.objects.filter(
            status__in=[SendQueue.STATUS_ERROR, SendQueue.STATUS_FAILED]
        ).count()
        pending = SendQueue.objects.filter(status=SendQueue.STATUS_PENDING).count()
        if stuck == 0:
            checks.append({'name': 'Очередь отправки', 'status': 'ok',
                            'detail': f'Ошибок нет · Ожидает: {pending}', 'icon': '📤'})
        else:
            checks.append({'name': 'Очередь отправки', 'status': 'warn',
                            'detail': f'Ошибок: {stuck} · Ожидает: {pending}', 'icon': '📤'})
    except Exception as e:
        checks.append({'name': 'Очередь отправки', 'status': 'error',
                        'detail': str(e), 'icon': '📤'})

    # Системная информация
    sys_info = {
        'python':   sys.version.split()[0],
        'platform': platform.system() + ' ' + platform.release(),
        'django':   __import__('django').get_version(),
        'db_path':  str(__import__('django').conf.settings.DATABASES['default']['NAME']),
        'time':     timezone.now().strftime('%d.%m.%Y %H:%M:%S'),
    }

    overall = 'ok'
    if any(c['status'] == 'error' for c in checks):
        overall = 'error'
    elif any(c['status'] == 'warn' for c in checks):
        overall = 'warn'

    # JSON-режим для внешних мониторингов
    if request.GET.get('format') == 'json':
        return JsonResponse({
            'status': overall,
            'checks': checks,
            'sys': sys_info,
        })

    return render(request, 'edi/healthcheck.html', {
        'checks':   checks,
        'sys_info': sys_info,
        'overall':  overall,
    })


# ═══════════════════════════════════════════════════════
# Backup БД
# ═══════════════════════════════════════════════════════

def backup_db(request):
    """GET /backup/ — скачать резервную копию SQLite."""
    import shutil, io
    from django.conf import settings as djset

    db_path = djset.DATABASES['default']['NAME']
    buf = io.BytesIO()
    with open(db_path, 'rb') as f:
        buf.write(f.read())
    buf.seek(0)

    fname = f'docrobot_backup_{date.today().strftime("%Y%m%d_%H%M")}.sqlite3'
    ActivityLog.objects.create(level='info', action='db_backup', message=f'Скачана резервная копия БД: {fname}')

    response = HttpResponse(buf.read(), content_type='application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def api_cleanup_now(request):
    """POST /api/cleanup/ — ручной запуск очистки старых документов."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)

    import threading
    def _run():
        try:
            from django.core.management import call_command
            call_command('cleanup_old')
        except Exception as e:
            ActivityLog.objects.create(level='error', action='cleanup', message=str(e))

    threading.Thread(target=_run, daemon=True).start()
    return JsonResponse({'ok': True, 'message': 'Очистка запущена в фоне'})


# ═══════════════════════════════════════════════════════
# Статистика по поставщикам
# ═══════════════════════════════════════════════════════

def suppliers(request):
    """Аналитика по GLN поставщиков."""
    days = int(request.GET.get('days', 30))
    since = timezone.now() - timedelta(days=days)

    # Топ поставщиков по количеству
    top_by_count = list(
        EdiDocument.objects
        .filter(received_at__gte=since)
        .exclude(supplier_gln='')
        .values('supplier_gln', 'supplier_name')
        .annotate(
            total=Count('id'),
            orders=Count('id', filter=Q(doc_type='ORDER')),
            invoices=Count('id', filter=Q(doc_type='INVOICE')),
            desadv=Count('id', filter=Q(doc_type='DESADV')),
        )
        .order_by('-total')[:20]
    )

    # Динамика по дням для топ-5 поставщиков
    top5_glns = [s['supplier_gln'] for s in top_by_count[:5]]
    daily_by_supplier = {}
    for gln in top5_glns:
        name = next((s['supplier_name'] or gln for s in top_by_count if s['supplier_gln'] == gln), gln)
        points = []
        for i in range(min(days, 30) - 1, -1, -1):
            day = date.today() - timedelta(days=i)
            cnt = EdiDocument.objects.filter(
                supplier_gln=gln, received_at__date=day
            ).count()
            points.append({'day': day.strftime('%d.%m'), 'count': cnt})
        daily_by_supplier[name[:20]] = points

    # Общие цифры за период
    total_docs      = EdiDocument.objects.filter(received_at__gte=since).count()
    unique_suppliers = EdiDocument.objects.filter(received_at__gte=since).exclude(supplier_gln='').values('supplier_gln').distinct().count()

    return render(request, 'edi/suppliers.html', {
        'days':              days,
        'periods':           [7, 14, 30, 90],
        'top_by_count':      top_by_count,
        'total_docs':        total_docs,
        'unique_suppliers':  unique_suppliers,
        'avg_per_supplier':  round(total_docs / unique_suppliers, 1) if unique_suppliers else 0,
        'daily_json':        json.dumps(daily_by_supplier),
        'top5_glns':         top5_glns,
    })


# ═══════════════════════════════════════════════════════
# Просмотр лог-файлов
# ═══════════════════════════════════════════════════════

def log_files(request):
    """GET /log-files/ — просмотр файловых логов прямо в браузере."""
    import os
    from django.conf import settings as djset

    logs_dir = getattr(djset, 'LOGS_DIR', djset.BASE_DIR / 'logs')
    selected = request.GET.get('file', 'docrobot.log')
    lines    = int(request.GET.get('lines', 200))

    # Список доступных файлов
    available = []
    try:
        for f in sorted(os.listdir(logs_dir)):
            if f.endswith('.log'):
                fpath = logs_dir / f
                size  = os.path.getsize(fpath)
                available.append({'name': f, 'size': size, 'size_kb': round(size / 1024, 1)})
    except FileNotFoundError:
        pass

    content = ''
    file_size = 0
    if available:
        # Проверяем что файл из разрешённого списка
        valid_names = [f['name'] for f in available]
        if selected not in valid_names and valid_names:
            selected = valid_names[0]

        fpath = logs_dir / selected
        try:
            file_size = os.path.getsize(fpath)
            with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
                all_lines = f.readlines()
            # Берём последние N строк
            content = ''.join(all_lines[-lines:])
        except FileNotFoundError:
            content = '(файл пуст или не существует)'
        except Exception as e:
            content = f'Ошибка чтения: {e}'

    return render(request, 'edi/log_files.html', {
        'available':  available,
        'selected':   selected,
        'content':    content,
        'lines':      lines,
        'file_size':  round(file_size / 1024, 1),
        'logs_dir':   str(logs_dir),
    })
